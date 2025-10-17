# src/scripts/integrate_player_summaries.py
#!/usr/bin/env python3
"""
Integrate Player Summary Excel files into master_player_stats.json
(No pandas required - uses openpyxl for Excel processing)
"""
import json
import re
from datetime import datetime
from pathlib import Path
try:
    from openpyxl import load_workbook
except ImportError:
    print("❌ openpyxl is not installed. Please run: pip install openpyxl")
    exit(1)

def integrate_player_summaries():
    base_dir = Path.cwd()
    
    # Specific location where your master file is
    master_file = base_dir / "data" / "processed" / "master_stats.json"
    
    if not master_file.exists():
        print(f"❌ master_stats.json not found at: {master_file}")
        print("Please make sure the file exists at data/processed/master_stats.json")
        return False
    
    print(f"✅ Found master data at: {master_file}")
    
    # Load master data
    try:
        with open(master_file, 'r') as f:
            master_data = json.load(f)
        print(f"✅ Loaded master data with {len(master_data)} players")
    except Exception as e:
        print(f"❌ Error reading master data: {e}")
        return False
    
    # Search for player Excel files in common locations
    search_dirs = [
        base_dir / "attached_assets",  # Most likely location
        base_dir,
        base_dir / "data" / "raw",
        base_dir / "assets"
    ]
    
    player_files = []
    for search_dir in search_dirs:
        if search_dir.exists():
            for pattern in ["*.xlsx", "*.xls"]:
                found_files = list(search_dir.glob(pattern))
                player_files.extend(found_files)
                if found_files:
                    print(f"✅ Found {len(found_files)} Excel files in {search_dir}")
    
    # Remove duplicates
    player_files = list(set(player_files))
    
    print(f"🔍 Total player summary files found: {len(player_files)}")
    
    if not player_files:
        print("❌ No Excel files found. Please check file locations.")
        print("Expected to find Excel files in: attached_assets/, data/raw/, or assets/")
        return False
    
    integrated_count = 0
    
    for player_file in player_files:
        try:
            print(f"\n📊 Processing {player_file.name}...")
            
            # Extract player name from filename
            player_name = extract_player_name(player_file.name)
            
            if not player_name:
                print(f"⚠️  Could not extract player name from {player_file.name}")
                continue
            
            # Find matching player in master data
            player_id, player_data = find_player_in_master(player_name, master_data)
            
            if not player_id:
                print(f"⚠️  Player '{player_name}' not found in master data")
                continue
            
            print(f"✅ Found match: {player_data.get('name')} ({player_data.get('team')})")
            
            # Process the Excel file
            player_summary_data = process_player_excel(player_file)
            
            if player_summary_data:
                # Integrate the data
                if 'historical_analysis' not in player_data:
                    player_data['historical_analysis'] = {}
                
                player_data['historical_analysis'].update(player_summary_data)
                player_data['historical_analysis']['last_updated'] = datetime.now().isoformat()
                player_data['historical_analysis']['source_file'] = player_file.name
                
                integrated_count += 1
                print(f"✅ Integrated historical data for {player_name}")
            
        except Exception as e:
            print(f"❌ Error processing {player_file.name}: {e}")
            continue
    
    # Save updated master data
    try:
        with open(master_file, 'w') as f:
            json.dump(master_data, f, indent=2)
        print(f"\n💾 Saved updated master data to {master_file}")
    except Exception as e:
        print(f"❌ Error saving updated master data: {e}")
        return False
    
    print(f"\n🎉 Player summary integration complete!")
    print(f"   Integrated data for {integrated_count} players")
    print(f"   Total players in master data: {len(master_data)}")
    
    return True

def extract_player_name(filename):
    """Extract player name from filename"""
    # Remove file extension and any numbers/special characters
    name = re.sub(r'\.xlsx$|\.xls$', '', filename, flags=re.IGNORECASE)
    name = re.sub(r'[_\-]\d+$', '', name)
    name = re.sub(r'_\d+\.\d+\.\d+$', '', name)
    name = re.sub(r'_\d+$', '', name)  # Remove trailing numbers
    
    # Clean up any remaining special characters
    name = re.sub(r'[^a-zA-Z\s]', ' ', name)
    name = ' '.join(name.split())  # Remove extra spaces
    
    return name.strip()

def find_player_in_master(player_name, master_data):
    """Find player in master data using fuzzy matching"""
    # Clean the search name
    search_name = player_name.lower().strip()
    
    # First try exact match
    for player_id, player_data in master_data.items():
        master_name = player_data.get('name', '').lower()
        if master_name == search_name:
            return player_id, player_data
    
    # Try partial matches (player name contained in master name or vice versa)
    for player_id, player_data in master_data.items():
        master_name = player_data.get('name', '').lower()
        if search_name in master_name or master_name in search_name:
            return player_id, player_data
    
    # Try matching by first and last name
    search_parts = search_name.split()
    if len(search_parts) >= 2:
        first_name, last_name = search_parts[0], search_parts[-1]
        
        for player_id, player_data in master_data.items():
            master_name = player_data.get('name', '').lower()
            master_parts = master_name.split()
            
            if (len(master_parts) >= 2 and 
                first_name == master_parts[0].lower() and 
                last_name == master_parts[-1].lower()):
                return player_id, player_data
    
    # Try matching just by last name
    if len(search_parts) >= 1:
        last_name = search_parts[-1]
        for player_id, player_data in master_data.items():
            master_name = player_data.get('name', '').lower()
            master_parts = master_name.split()
            
            if master_parts and master_parts[-1].lower() == last_name:
                return player_id, player_data
    
    return None, None

def process_player_excel(file_path):
    """Process player Excel file and extract valuable data"""
    try:
        workbook = load_workbook(filename=file_path, data_only=True)
        
        result = {
            'career_averages': {},
            'opponent_splits': {},
            'game_logs': [],
            'advanced_metrics': {}
        }
        
        # Process Career Averages sheet
        if 'Career Averages' in workbook.sheetnames:
            sheet = workbook['Career Averages']
            data = sheet_to_dict(sheet)
            
            for row in data:
                if row.get('YR') and isinstance(row.get('YR'), (int, float)):
                    year = int(row['YR'])
                    result['career_averages'][str(year)] = {
                        'games_played': safe_float(row.get('GM')),
                        'fantasy_points_avg': safe_float(row.get('FP')),
                        'supercoach_avg': safe_float(row.get('SC')),
                        'kicks_avg': safe_float(row.get('K')),
                        'handballs_avg': safe_float(row.get('H')),
                        'marks_avg': safe_float(row.get('M')),
                        'tackles_avg': safe_float(row.get('T')),
                        'goals_avg': safe_float(row.get('G')),
                        'behinds_avg': safe_float(row.get('B')),
                        'time_on_ground_avg': safe_float(row.get('TOG')),
                        'disposal_efficiency': safe_float(row.get('DE%'))
                    }
        
        # Process Opponent Splits sheet
        if 'Opponent Splits' in workbook.sheetnames:
            sheet = workbook['Opponent Splits']
            data = sheet_to_dict(sheet)
            
            for row in data:
                opponent = row.get('OPP')
                if opponent and opponent != 'OPP' and opponent != 'Team':
                    result['opponent_splits'][opponent] = {
                        'games_played': safe_float(row.get('GM')),
                        'fantasy_points_avg': safe_float(row.get('FP')),
                        'supercoach_avg': safe_float(row.get('SC')),
                        'kicks_avg': safe_float(row.get('K')),
                        'handballs_avg': safe_float(row.get('H')),
                        'marks_avg': safe_float(row.get('M')),
                        'tackles_avg': safe_float(row.get('T')),
                        'goals_avg': safe_float(row.get('G')),
                        'behinds_avg': safe_float(row.get('B'))
                    }
        
        # Process Game Logs sheet
        if 'Game Logs' in workbook.sheetnames:
            sheet = workbook['Game Logs']
            data = sheet_to_dict(sheet)
            
            for row in data:
                if row.get('YR') and row.get('RD'):
                    game_log = {
                        'year': int(row['YR']) if safe_float(row.get('YR')) else None,
                        'round': str(row['RD']),
                        'opponent': row.get('OPP'),
                        'venue': row.get('VEN'),
                        'fantasy_points': safe_float(row.get('FP')),
                        'supercoach_score': safe_float(row.get('SC')),
                        'kicks': safe_float(row.get('K')),
                        'handballs': safe_float(row.get('H')),
                        'marks': safe_float(row.get('M')),
                        'tackles': safe_float(row.get('T')),
                        'goals': safe_float(row.get('G')),
                        'behinds': safe_float(row.get('B')),
                        'time_on_ground': safe_float(row.get('TOG')),
                        'disposal_efficiency': safe_float(row.get('DE%'))
                    }
                    # Only add if we have basic data
                    if game_log['fantasy_points'] is not None:
                        result['game_logs'].append(game_log)
        
        # Calculate advanced metrics
        result['advanced_metrics'] = calculate_advanced_metrics(result)
        
        print(f"   Extracted: {len(result['game_logs'])} game logs, {len(result['career_averages'])} seasons, {len(result['opponent_splits'])} opponents")
        return result
        
    except Exception as e:
        print(f"❌ Error processing Excel file {file_path.name}: {e}")
        return None

def sheet_to_dict(sheet):
    """Convert Excel sheet to list of dictionaries"""
    data = []
    headers = []
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if row_idx == 0:
            # First row is headers
            headers = [str(cell).strip() if cell is not None else f"col_{i}" for i, cell in enumerate(row)]
        else:
            # Data rows
            row_data = {}
            for col_idx, cell in enumerate(row):
                if col_idx < len(headers):
                    row_data[headers[col_idx]] = cell
            # Only add row if it has data
            if any(cell is not None for cell in row):
                data.append(row_data)
    
    return data

def safe_float(value):
    """Safely convert value to float, handling None and strings"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Try to extract numbers from strings like "5.5 (60%)"
        numeric_part = value.split()[0] if ' ' in value else value
        try:
            return float(numeric_part)
        except ValueError:
            return None
    return None

def calculate_advanced_metrics(player_data):
    """Calculate advanced metrics from player data"""
    metrics = {
        'consistency_rating': 0,
        'best_matchup': None,
        'worst_matchup': None
    }
    
    # Calculate consistency from game logs
    fantasy_scores = [log['fantasy_points'] for log in player_data.get('game_logs', []) 
                     if log.get('fantasy_points') is not None]
    
    if fantasy_scores and len(fantasy_scores) > 1:
        avg_score = sum(fantasy_scores) / len(fantasy_scores)
        std_dev = (sum((x - avg_score) ** 2 for x in fantasy_scores) / len(fantasy_scores)) ** 0.5
        metrics['consistency_rating'] = round(avg_score / std_dev, 2) if std_dev > 0 else 0
    
    # Analyze opponent splits for strengths/weaknesses
    opponent_data = player_data.get('opponent_splits', {})
    if opponent_data:
        valid_opponents = [(opp, data.get('fantasy_points_avg')) 
                          for opp, data in opponent_data.items() 
                          if data.get('fantasy_points_avg') is not None]
        
        if valid_opponents:
            valid_opponents.sort(key=lambda x: x[1], reverse=True)
            
            metrics['best_matchup'] = {
                'opponent': valid_opponents[0][0],
                'average_score': valid_opponents[0][1]
            }
            
            if len(valid_opponents) > 1:
                metrics['worst_matchup'] = {
                    'opponent': valid_opponents[-1][0],
                    'average_score': valid_opponents[-1][1]
                }
    
    return metrics

if __name__ == "__main__":
    integrate_player_summaries()